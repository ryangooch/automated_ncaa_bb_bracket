# -*- coding: utf-8 -*-

"""
This is experimental development to extend the functionality of the automated
bracket code that I've been working on for a few years.

I have written a set of utilities to scrape actual raw ratings data from a few
computer metrics. Currently supported are:

    * Kenpom
    * ESPN BPI
    * Dokter Entropy

The scripts produce csv files with the names

|     Metric     |       CSV File      |
| -------------- | ------------------- |
|     Kenpom     | kenpom_YYYYMMDD.csv |
| Dokter Entropy | dokent_YYYYMMDD.csv |
|    ESPN BPI    | bpi_YYYYMMDD.csv    |

The goal is to begin to incorporate this as a separate class. Eventually all
will be refactored into a coherent project structure. For now, goal is to get 
a working automated bracket.

I am essentially starting from the code in bracket_picker.py and editing here.
"""

from urllib.request import urlretrieve
from openpyxl import Workbook, load_workbook
from scrape import download_kenpom, download_dokent, download_bpi

import numpy as np
import pandas as pd
import csv
import datetime
import warnings
import os

warnings.filterwarnings('ignore')

def rank_calc(x, y) :
    """
    Calculates the final ranking for teams. 
    Inputs
    x:  Computer ranking
    y:  Human Ranking
    Outputs
    If both rankings present,
    Rank = (2 * x + y) / 3
    If only computer ranking present, returns computer ranking.
    If user prefers another formula, it must be in this format
    """
    if np.isnan(y) :
        return x
    else:
        return ((3 * x + y) / 4.)

def remove_seed(s):
    # this is needed to clean up kenpom names after tourney begins
    return (''.join([i for i in s if not i.isdigit()])).strip()

class Bracketeer(object):
    """
    Downloads, aggregates data, selects final 68 teams for NCAA 
    Tournament following a simple average of computer and human
    rankings.
    """
    def __init__(self, csv_save_path = 'masseyratings.csv'):
        self.save_path = csv_save_path
        

        # List for seeding teams
        self.seeds = np.array([
            1,1,1,1,
            2,2,2,2,
            3,3,3,3,
            4,4,4,4,
            5,5,5,5,
            6,6,6,6,
            7,7,7,7,
            8,8,8,8,
            9,9,9,9,
            10,10,10,10,
            11,11,11,11,11,11,
            12,12,12,12,
            13,13,13,13,
            14,14,14,14,
            15,15,15,15,
            16,16,16,16,16,16
        ])

        self.download_csv()
        self.parse_csv()

        
    def download_csv(self) :
        """
        Get the composite csv from masseyratings.com
        stores in a folder, overwrites existing unless
        name specified
        """
        composite_csv = 'http://www.masseyratings.com/cb/compare.csv'
        urlretrieve(composite_csv, self.save_path)

    def parse_csv(self) :
        """
        need to parse the csv
        first row of interest starts at "team" in column 1
        data starts two rows after
        """
        data = []
        with open(self.save_path, newline='') as csvfile:
            reader = csv.reader(csvfile, delimiter = ',')
            for row in reader:
                data.append(row)

        # Input Sanitization
        # The csv file wasn't formatted in a way pandas could read. 
        # There is a header section followed by the data, but the header
        # section is variable in length. We can simply iterate through 
        # until we see the 'Team' string indicating the start of the data

        i = 0 # counter of lines until Team found
        for row in data:
            if row and row[0] == 'Team':
                break
            i = i + 1

        # The header data contains the abbreviations and URLs for the
        # various ranking systems in the dataset, might be useful later
        self.header_data = data[:i]
        team_data = data[i:]
        column_names = team_data[0] # Dataset column names
        # strip whitespace from the column names for normalization
        for i in range(len(column_names)):
            column_names[i] = column_names[i].lstrip()

        # drop the data set in a pandas Dataframe
        self.team_data_df = pd.DataFrame(
            team_data[2:],
            columns=column_names
        )
        
        # remove team name whitespace
        self.team_data_df = self.team_data_df.apply(
            lambda x: x.str.strip() if x.dtype == "object" else x)

    def print_polls(self) :
        """
        Prints available polls for convenience
        """
        print(self.team_data_df.columns)
        
    def get_conferences(self):
        return pd.unique(self.team_data_df['Conf'])
    
    def get_tourney_teams (self, comp_polls = None, rank_calc_func = None,
            conf_winners = None, use_metrics = False, human_polls = True) :
        """
        Analysis on the full dataset to derive the teams actually in the
        tournament

        Inputs:
            comp_polls: List of strings representing subset of available polls
                to pull from for computer rankings
            rank_calc_func: Function to use for calculation weighted average
                of computer and human polls for final rank calculation. If none,
                default is used (weight for computer polls to human polls 3:1)
            conf_winners: Dictionary where keys are conferences and values are
                teams who have won the autobid for the specified conferences
            use_metrics: Boolean. If true, attempts to pull data from csv files
                representing the raw rating data for polls in comp_polls. If
                false, uses rank data to aggregate
            human_polls: Boolean. If true, uses human polls in final 
                computation. If false, ignores human polls
        """

        # idea here: splitting off functionality to be more modular, but I want
        # access to these in a few places. May delete some if needed
        self.comp_polls = comp_polls # default: None
        self.rank_calc_func = rank_calc_func # default: None
        self.conf_winners = conf_winners # default: None
        self.use_metrics = use_metrics # default: False
        self.human_polls = human_polls # default: True
        
        # Use a place holder dataframe for calculated means and ranks 
        # for all teams
        summary_df = self.team_data_df[["Team","Conf"]]

        if use_metrics is False:
            try: 
                # returns dataframes with mean in df['mean']
                comp_rankings = self.get_comp_rankings()
                summary_df["comp_mean"] = comp_rankings['mean']

            except Exception as e:
                print('Exception occured:')
                print('------------------')
                raise e

        elif use_metrics is True:
            try:
                comp_ratings = self.get_comp_ratings()
                summary_df["comp_mean"] = comp_ratings['mean']

            except Exception as e:
                print('Exception occured:')
                print('------------------')
                raise e
        else:
            print('use_metrics must be True or False')
            raise Exception('use_metrics must be True or False. Input was {}'\
                .format(use_metrics)) 

        if use_metrics is False and human_polls is True:
            # basically right now I have no intention of including human ranks
            # and then averaging with ranks from mean computer ratings. May
            # change in future
            try: 
                human_rankings = self.get_human_rankings()
                summary_df["human_mean"] = human_rankings['mean']

            except Exception as e:
                print('Exception occured:')
                print('------------------')
                raise e
        else:
            # this will essentially make the final rankings calculation
            # ignore any human poll inputs. Initializing a human rankings
            # 'mean' filled with nans
            human_rankings = pd.DataFrame(
                data = {
                    'mean':np.full([self.team_data_df.shape[0],],np.nan)
                },
                columns = ['mean']
            )
            summary_df["human_mean"] = human_rankings['mean']

        # Calculate the final rank using either the user-defined algorithm
        # or the default
        if rank_calc_func is None :
            rank_calc_func = rank_calc

        summary_df["final_rank"] = np.vectorize(rank_calc_func)(\
                summary_df["comp_mean"], summary_df["human_mean"])

        # Sort by calculated rank
        if use_metrics is True:
            summary_df.sort_values(by=['final_rank'],inplace=True,ascending=False)
        else:
            summary_df.sort_values(by=['final_rank'],inplace=True)

        # ----
        # Tourney rules dictate the winners of the conferences all have auto
        # bids into the tourney. There are 32 of these. The remaining 36
        # are chosen at large. We simply take the highest ranked team in each
        # conference as an auto bid. For the final bracket, we accept 
        # a list of conference winners
        # ----
        # now that final rank is calculated, need to find auto-bids and 
        # at large
        
        auto_bid_confs = pd.unique(summary_df['Conf'])

        # Using groupby to grab the auto bids
        auto_bid_teams = summary_df.groupby(['Conf']).head(1)
            
        if conf_winners is not None:
            auto_bid_teams = self._replace_auto_bid(conf_winners,auto_bid_teams)
        else:
            auto_bid_teams = auto_bid_teams['Team'].values
        # slated for removal below:
#         else :
#             auto_bid_confs = pd.unique(summary_df['Conf'])

#             # Using groupby to grab the auto bids
#             auto_bid_teams = summary_df.groupby(['Conf']).head(1)['Team'].values

        # and we can use ~isin now to get at larges
        at_large_teams = summary_df[~summary_df['Team'].isin(auto_bid_teams)].head(36)['Team'].values

        # all 68 teams in one array
        all_68 = np.append(auto_bid_teams,at_large_teams)
        
        self.auto_bid_teams = auto_bid_teams
        self.at_large_teams = at_large_teams
        self.all_68 = all_68

        # First four next four
        self._ffnf = summary_df[~summary_df['Team'].isin(auto_bid_teams)].iloc[36:44]['Team'].values

        self.final_68 = summary_df[summary_df['Team'].isin(all_68)]

        self.final_68["seed"] = self.seeds

        # return the dataframe to the user
        # return self.final_68

    def get_comp_rankings(self):
        """
        return computer rankings dataframe
        """
        # Drop columns that are unnecessary 
        cols_to_drop = ["WL","Rank","Mean","Trimmed","Median","StDev","AP",
            "USA"]

        # create computer ranking dataframe
        comp_rankings = self.team_data_df.drop(cols_to_drop,axis=1)

        # If user provides list of specific computer polls to use, subset here
        # The abbreviations will need to be used. If one desired poll isn't
        # available, then it will return a KeyError. In this case, let user 
        # know, then break out after printing all available polls.
        if self.comp_polls is not None :
            try:
                comp_rankings = comp_rankings[self.comp_polls]
            except KeyError as e:
                print('One or more of the polls you tried isn\'t available\n')
                print(comp_rankings.columns)
                raise e

        # Convert computer rankings dataframe columns to numeric where
        # appropriate
        comp_rankings = comp_rankings.apply(\
            lambda x: pd.to_numeric(x, errors='ignore'))

        # compute arithmetic mean of computer rankings
        comp_rankings['mean'] = comp_rankings.mean(axis=1,numeric_only=True)

        return comp_rankings

    def get_human_rankings(self):
        """
        return human rankings from dataframe
        """

        # create human ranking dataframe
        try:
            human_rankings = self.team_data_df[['Team','Conf','AP','USA']]
        except KeyError as e:
            print('Human rankings are unavailable. Change human_polls to False')
            print('if desired')
            raise e
        
        # remove whitespace around strings in these polls
        human_rankings['AP'] = human_rankings['AP'].str.strip()
        human_rankings['USA'] = human_rankings['USA'].str.strip()

        # convert rankings to numeric
        human_rankings = human_rankings.apply(\
            lambda x: pd.to_numeric(x, errors='ignore'))

        human_rankings['mean'] = human_rankings.mean(
            axis = 1, numeric_only = True, skipna = True)

        return human_rankings

    def get_comp_ratings(self):
        """
        return raw rating data instead of rankings for specified columns

        right now only three available, so will basically ignore list of inputs
        for now
        """

        # Download latest data. This includes some waiting to be respectful to
        # servers, so display message to let user know it will take a second
        # all files downloaded to csv_files/

        # for now, to save time re-running the code, including logic to only
        # download if file is not there.
        print('Downloading Ratings Data')

        if not os.path.isfile('csv_files/bpi.csv'):
            download_bpi()
        
        if not os.path.isfile('csv_files/dokent.csv'):
            download_dokent()
        
        if not os.path.isfile('csv_files/kenpom.csv'):
            download_kenpom()
        
        print('Downloading Finished!')

        kenpom_df = pd.read_csv('csv_files/kenpom.csv',index_col=False)
        bpi_df = pd.read_csv('csv_files/bpi.csv',index_col=False)
        dokent_df = pd.read_csv('csv_files/dokent.csv',index_col=False)
        
        # remove seed from kenpom
        kenpom_df['Team'] = kenpom_df['Team'].map(remove_seed)
        
        # one big issue is that the names don't match. Ken Massey uses a
        # database of keys and values to massage the names into a consistent
        # format. I've implemented this similarly, using distinct csv files
        # for every translation. I'm using massey names as my standard, and 
        # converting the rest accordingly. The keys will be stored in 
        # csv_files for now, and conversion will happen here. This all may be
        # moved to makefiles or something down the road

        massey_to_bpi = pd.read_csv('csv_files/names_massey-bpi.csv')
        massey_to_dokter = pd.read_csv('csv_files/names_massey-dokter.csv')
        massey_to_kenpom = pd.read_csv('csv_files/names_massey-kenpom.csv')

        for row in zip(massey_to_kenpom['Massey'].values,
            massey_to_kenpom['Kenpom'].values):
            
            massey_team = row[0]
            kenpom_team = row[1]
            kenpom_df.loc[kenpom_df['Team'].str.match("^%s$"%kenpom_team), \
                'Team'] = massey_team
        
        for row in zip(massey_to_bpi['Massey'].values,
            massey_to_bpi['BPI'].values):
            
            massey_team = row[0]
            bpi_team = row[1]
            bpi_df.loc[bpi_df['Team'].str.match("^%s$"%bpi_team), \
                'Team'] = massey_team

        for row in zip(massey_to_dokter['Massey'].values,
            massey_to_dokter['Dokter'].values):
            
            massey_team = row[0]
            dokent_team = row[1]
            dokent_df.loc[dokent_df['Team'].str.match("^%s$"%dokent_team), \
                'Team'] = massey_team

        # drop unnecessary columns
        dokent_df.drop(columns=['Unnamed: 0','Rk'],inplace=True)
        bpi_df.drop(columns=['Unnamed: 0','Rk','Conf','W-L'],inplace=True)
        kenpom_df.drop(columns=['Unnamed: 0','Rk','W-L'],inplace=True)

        # print(self.team_data_df.columns)
        # print(dokent_df.columns)
        # print(kenpom_df.columns)
        # print(bpi_df.columns)

        comp_ratings = self.team_data_df.merge(
            right=dokent_df,
            on='Team',
            how='inner',
            suffixes=('','_dokter'),
            validate='one_to_one'
        )

        comp_ratings = comp_ratings.merge(
            right=kenpom_df,
            on='Team',
            how='inner',
            suffixes=('','_kenpom'),
            validate='one_to_one'
        )

        comp_ratings = comp_ratings.merge(
            right=bpi_df,
            on='Team',
            how='inner',
            suffixes=('','_bpi'),
            validate='one_to_one'
        )

        # now standardize the summary rankings for each of the above. This is
        # acceptable I think, because the generating functions in each case
        # are designed to generate normal data, though the mean and variance
        # are all different. I standardize to mean 0 and variance 1, then take
        # the mean.
        comp_ratings['BPI'] = (comp_ratings['BPI'] - \
            comp_ratings['BPI'].mean()) / comp_ratings['BPI'].std()
        comp_ratings['AdjEM'] = (comp_ratings['AdjEM'] - \
            comp_ratings['AdjEM'].mean()) / comp_ratings['AdjEM'].std()
        comp_ratings['power'] = (comp_ratings['power'] - \
            comp_ratings['power'].mean()) / comp_ratings['power'].std()

        comp_ratings['mean'] = \
            comp_ratings[['BPI','AdjEM','power']].mean(axis=1)
        
        return comp_ratings
        
    def _replace_auto_bid(self,conf_winners,auto_bid_teams):
        """
        Replaces auto bid assumptions with actual winners. Winners expected
        to be in dict where keys are conference names that are consistent
        with internal names, while auto_bid_teams is a dataframe where
        the top ranked team in each conference is present. Returns a list
        of updated auto bid teams.
        """
        try:
            for conf, team in conf_winners.items():
                if team is not None:
                    auto_bid_teams.loc[auto_bid_teams['Conf']==conf, 'Team'] = team
            
        except Exception as e:
            print('it failed')
            raise(e)
            
        return auto_bid_teams['Team'].values

    def fill_bracket(self) :
        # Need to tell Excel which cells get which team. Because we're 
        # using 'snake' method, there are specific cells corresponding
        # to each seed and rank. Easiest way to do this is to simply
        # hard code the table in, for now
        excel_placements = [
            'C7', 'R7', 'R39', 'C39', 'C67', 'R67', 'R35', 'C35', 
            'C27', 'R27', 'R59', 'C59', 'C51', 'R51', 'R19', 'C19',
            'C15', 'R15', 'R47', 'C47', 'C55', 'R55', 'R23', 'C23',
            'C31', 'R31', 'R63', 'C63', 'C43', 'R43', 'R11', 'C11',
            'C13', 'R13', 'R45', 'C45', 'C65', 'R65', 'R33', 'C33',
            'C25', 'R25', 'Q71', 'Q73', 'D71', 'D73', # 11 seeds
            'C49', 'R49', 'R17', 'C17', 'C21', 'R21', 'R53', 'C53',
            'C61', 'R61', 'R29', 'C29', 'C37', 'R37', 'R69', 'C69',
            'C41', 'R41', 'O71', 'O73', 'F71', 'F73'
        ]

        # append that to our final_68 dataframe
        self.final_68['excel'] = excel_placements

        # input to the workbook. There is one set up called 'bracket.xlsx'
        wb = Workbook()
        ws = wb.active
        for index, row in self.final_68.iterrows():
            ws[row['excel']] = row['Team']

        # get today's date, save bracket
        today = str(datetime.date.today())
        save_file = 'bracket' + today + '.xlsx'
        wb.save(save_file)