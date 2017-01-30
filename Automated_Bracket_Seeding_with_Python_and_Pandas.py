
# coding: utf-8

# ## Automated NCAA Tournament Bracket Filling with Python and Pandas

from urllib.request import urlretrieve
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
import csv
import datetime

import warnings
warnings.filterwarnings('ignore')

def rank_calc(x, y) :
    """
    Calculates the final ranking for teams. 
    Inputs
    x:  Computer ranking
    y:  Human Ranking
    Outputs
    If both rankings present,

    Rank = (2 * x + y) / 3

    If only computer ranking present, returns computer ranking.

    If user prefers another formula, it must be in this format
    """
    if np.isnan(y) :
        return x
    else:
        return ((2 * x + y) / 3.)

class Bracketeer(object):
    """
    Downloads, aggregates data, selects final 68 teams for NCAA 
    Tournament following a simple average of computer and human
    rankings.
    """
    def __init__(self, csv_save_path = 'masseyratings.csv'):
        self.save_path = csv_save_path
        

        # List for seeding teams
        self.seeds = np.array([
            1,1,1,1,
            2,2,2,2,
            3,3,3,3,
            4,4,4,4,
            5,5,5,5,
            6,6,6,6,
            7,7,7,7,
            8,8,8,8,
            9,9,9,9,
            10,10,10,10,
            11,11,11,11,11,11,
            12,12,12,12,
            13,13,13,13,
            14,14,14,14,
            15,15,15,15,
            16,16,16,16,16,16
        ])

        self.download_csv()
        self.parse_csv()

        
    def download_csv(self) :
        """
        Get the composite csv from masseyratings.com
        stores in a folder, overwrites existing unless
        name specified
        """
        composite_csv = 'http://www.masseyratings.com/cb/compare.csv'
        urlretrieve(composite_csv, self.save_path)

    def parse_csv(self) :
        """
        need to parse the csv
        first row of interest starts at "team" in column 1
        data starts two rows after
        """
        data = []
        with open(self.save_path, newline='') as csvfile:
            reader = csv.reader(csvfile, delimiter = ',')
            for row in reader:
                data.append(row)

        # Input Sanitization
        # The csv file wasn't formatted in a way pandas could read. 
        # There is a header section followed by the data, but the header
        # section is variable in length. We can simply iterate through 
        # until we see the 'Team' string indicating the start of the data

        i = 0 # counter of lines until Team found
        for row in data:
            if row and row[0] == 'Team':
                break
            i = i + 1

        # The header data contains the abbreviations and URLs for the
        # various ranking systems in the dataset, might be useful later
        self.header_data = data[:i]
        team_data = data[i:]
        column_names = team_data[0] # Dataset column names
        # strip whitespace from the column names for normalization
        for i in range(len(column_names)):
            column_names[i] = column_names[i].lstrip()

        # drop the data set in a pandas Dataframe
        self.team_data_df = pd.DataFrame(
            team_data[2:],
            columns=column_names
        )

    def print_polls(self) :
        """
        Prints available polls for convenience
        """
        print(self.team_data_df.columns)
    def get_tourney_teams (self, comp_polls = None, rank_calc_func = None,
            conf_winners = None) :
        """
        Analysis on the full dataset to derive the teams actually in the
        tournament
        """

        # Drop columns that are unnecessary 
        cols_to_drop = ["WL","Rank","Mean","Trimmed",
                        "Median","StDev","AP","USA"]

        # Split off human ratings and computer ratings
        comp_ratings = self.team_data_df.drop(cols_to_drop,axis=1) # computer ratings
        human_ratings = self.team_data_df[['Team','Conf','AP','USA']] # pulls out human ratings

        # If user provides list of specific computer polls to use, subset here
        # The abbreviations will need to be used. If one desired poll isn't
        # available, then it will return a KeyError. In this case, let user know,
        # then break out after printing all available polls.
        if comp_polls is not None :
            try:
                comp_ratings = comp_ratings[comp_polls]
            except KeyError as e:
                print('One or more of the polls you tried isn\'t available\n')
                print(comp_ratings.columns)
                return

        # Use a place holder dataframe for calculated means and ranks 
        # for all teams
        summary_df = self.team_data_df[["Team","Conf"]]

        # Text processing for human ratings values
        human_ratings['AP'] = human_ratings['AP'].str.strip()
        human_ratings['USA'] = human_ratings['USA'].str.strip()

        # More sanitization. Convert ranks from strings to numeric so we 
        # can math them
        human_ratings = human_ratings.apply(lambda x: pd.to_numeric(x, errors='ignore'))
        comp_ratings = comp_ratings.apply(lambda x: pd.to_numeric(x, errors='ignore'))

        # Drop the mean ranks in the summary dataframe. In future, I may
        # add functionality here to allow a different aggregating mechanic
        # such as geometric mean, median, drop highest and lowest, etc
        summary_df["comp_mean"] = comp_ratings.mean(axis=1,numeric_only=True)
        summary_df["human_mean"] = human_ratings.mean(axis=1,numeric_only=True,skipna=True)

        # Calculate the final rank using either the user-defined algorithm
        # or the default
        if rank_calc_func is None :
            rank_calc_func = rank_calc

        summary_df["final_rank"] = np.vectorize(rank_calc_func)(\
                summary_df["comp_mean"], summary_df["human_mean"])

        # Sort by calculated rank
        summary_df.sort_values(by=['final_rank'],inplace=True)

        # ----
        # Tourney rules dictate the winners of the conferences all have auto
        # bids into the tourney. There are 32 of these. The remaining 36
        # are chosen at large. We simply take the highest ranked team in each
        # conference as an auto bid. For the final bracket, we accept 
        # a list of conference winners
        # ----
        # now that final rank is calculated, need to find auto-bids and 
        # at large
        if conf_winners is not None :
            auto_bid_teams = conf_winners
        else :
            auto_bid_confs = pd.unique(summary_df['Conf'])

            # Using groupby to grab the auto bids
            auto_bid_teams = summary_df.groupby(['Conf']).head(1)['Team'].values

        # and we can use ~isin now to get at larges
        at_large_teams = summary_df[~summary_df['Team'].isin(auto_bid_teams)].head(36)['Team'].values

        # all 68 teams in one array
        all_68 = np.append(auto_bid_teams,at_large_teams)

        self.final_68 = summary_df[summary_df['Team'].isin(all_68)]

        self.final_68["seed"] = self.seeds

        # return the dataframe to the user
        # return self.final_68

    def fill_bracket(self) :
        # Need to tell Excel which cells get which team. Because we're 
        # using 'snake' method, there are specific cells corresponding
        # to each seed and rank. Easiest way to do this is to simply
        # hard code the table in, for now
        excel_placements = [
            'C7', 'R7', 'R39', 'C39', 'C67', 'R67', 'R35', 'C35', 
            'C27', 'R27', 'R59', 'C59', 'C51', 'R51', 'R19', 'C19',
            'C15', 'R15', 'R47', 'C47', 'C55', 'R55', 'R23', 'C23',
            'C31', 'R31', 'R63', 'C63', 'C43', 'R43', 'R11', 'C11',
            'C13', 'R13', 'R45', 'C45', 'C65', 'R65', 'R33', 'C33',
            'C25', 'R25', 'Q71', 'Q73', 'D71', 'D73', # 11 seeds
            'C49', 'R49', 'R17', 'C17', 'C21', 'R21', 'R53', 'C53',
            'C61', 'R61', 'R29', 'C29', 'C37', 'R37', 'R69', 'C69',
            'C41', 'R41', 'O71', 'O73', 'F71', 'F73'
        ]

        # append that to our final_68 dataframe
        self.final_68['excel'] = excel_placements

        # input to the workbook. There is one set up called 'bracket.xlsx'
        wb = Workbook()
        ws = wb.active
        for index, row in self.final_68.iterrows():
            ws[row['excel']] = row['Team']

        # get today's date, save bracket
        today = str(datetime.date.today())
        save_file = 'bracket' + today + '.xlsx'
        wb.save(save_file)

    # def bracket_stats(self) :
    #     """
    #     Convenience method for user for a few commonly requested
    #     statistics from the final teams included
    #     """
    #     final_68.groupby(['Conf']).agg('count')['Team'].sort_values(ascending = False)

    #     # final_68[['Team','Conf']]
    #     final_68[final_68['Conf'].isin(['ACC'])]

